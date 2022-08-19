import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Cliffex-Lead\\Documents\\book1.xlsx")

sheet = book.active
# cell = sheet.cell(row=2, column=1)
# print(cell.value)
#
# sheet.cell(row=2,column=2).value = "piyush"
#
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet["A2"].value)
#
#
# print(sheet["B1"].value)
# print("**********************************************************************")
#
# # print(sheet.cell().value)



for i in range(1, sheet.max_row+1):
    print()

    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)


















# print("**********************************************************************")
# for j in range(1, sheet.max_column+1):
#     print(sheet.cell(row=1, column=j).value)




